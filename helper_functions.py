import os
from time import sleep
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut, GeocoderServiceError
import openpyxl
from openpyxl.styles import Border, Side, PatternFill


# Helper functions #
def clear_screen(sleep_time=1):
    sleep(sleep_time)
    os.system('cls' if os.name == 'nt' else 'clear')


def wait_for_keypress(sleep_time=0):
    while True:
        input('\nPress "Enter" to continue...')
        sleep(sleep_time)
        break

def clear_or_create_sheet(workbook, sheet_name):
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet.delete_rows(1, sheet.max_row)
    else:
        sheet = workbook.create_sheet(title=sheet_name)
    return sheet


def create_excel_headers(sheet):
    headers = ["Fetched", "Latitude", "Longitude", "Date", "Hour",
               "Temperature", "Precipitation Category", "Precipitation in mm"]

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    grey_fill = PatternFill(start_color='00CCCCCC',
                            end_color='00CCCCCC', fill_type='solid')

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.border = thin_border
        cell.fill = grey_fill
        sheet.column_dimensions[openpyxl.utils.get_column_letter(
            col)].width = 20


def get_coordinates():
    geolocator = Nominatim(user_agent='GeocodingApp')
    while True:
        user_input = input(
            'Enter city name(type "exit" to exit application): ')
        if user_input.lower() == 'exit':
            print('\nAction cancelled.')
            clear_screen()
            return None

        try:
            location = geolocator.geocode(user_input)

            latitude = None
            longitude = None

            if location:
                latitude, longitude = round(
                    location.latitude, 6), round(location.longitude, 6)
                return latitude, longitude, location[0].split(',')[0]
            else:
                print('Geocoding failed. Check your input or try a different location.')
        except GeocoderTimedOut:
            print("Geocoder service timed out. Please try again.")
        except GeocoderServiceError:
            print("Geocoder service error. Please try again later.")
        except Exception as e:
            print(f"An unexpected error occurred: {e}")
