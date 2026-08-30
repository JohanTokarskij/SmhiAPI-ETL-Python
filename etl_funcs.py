import datetime
import os
import requests
from helper_funcs import create_excel_headers, clear_or_create_sheet, wait_for_keypress
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut, GeocoderServiceError
import openpyxl

EXCEL_FILE = 'Weather_dashboard.xlsx'
# The snow1g API reports 'predominant_precipitation_type_at_surface' using the
# WMO/GRIB2 code table 4.201, which replaces the 0-6 'pcat' scale of pmp3g v2.
PRECIPITATION_CATEGORIES = {0: 'No precipitation',
                            1: 'Rain',
                            2: 'Thunderstorm',
                            3: 'Freezing rain',
                            4: 'Mixed/ice',
                            5: 'Snow',
                            6: 'Wet snow',
                            7: 'Rain and snow',
                            8: 'Ice pellets',
                            9: 'Graupel',
                            10: 'Hail',
                            11: 'Drizzle',
                            12: 'Freezing drizzle',
                            13: 'Hail (less than 5 mm)',
                            14: 'Hail (5 mm or more)',
                            255: 'Missing'}

# MENU: 1.Add new location to Excel-dashboard #
# Extract data from SMHI Api #
def extract_smhi_data(latitude, longitude):
    """
    Fetches data from the SMHI snow1g (version 1) forecast API for a given longitude and latitude.

    Errors encountered during the API request or data processing are printed to the console.

    Returns:
        dict: The response from the API in JSON format, or None if an error occurs.
    """

    try:
        response = requests.get(
            f'https://opendata-download-metfcst.smhi.se/api/category/snow1g/version/1/geotype/point/lon/{longitude}/lat/{latitude}/data.json')

        if response.status_code == 404:
            print('The provided location is outside the valid geographic area of SMHI.')
            return None        
        elif response.status_code != 200:
            print(f'Error fetching data: HTTP status code {response.status_code}')
            return None
        
        return response.json()

    except requests.RequestException as e:
        print(f'Error during API request: {e}')
        return None
    except ValueError:
        raise
    except Exception as e:
        print(f'An unexpected error occurred: {e}')
        return None

# Transform data #
def transform_smhi_data(data, latitude, longitude):
    """
    Transforms and returns fetched weather data into a structured format.

    Parameters:
    data (dict): A dictionary containing weather data from SMHI.
    latitude (float): Latitude of the observation location (in decimal degrees).
    longitude (float): Longitude of the observation location (in decimal degrees).

    Returns:
    List[List[Union[str, int, float]]]: A list of weather data where each sublist contains:
        0. 'fetched': Timestamp of when the data processing occurred ('YYYY-MM-DD hh:mm').
        1. 'longitude': Longitude of the observation location (in decimal degrees).
        2. 'latitude': Latitude of the observation location (in decimal degrees).
        3. 'date': Date of the weather data ('YYYY-MM-DD').
        4. 'hour': Hour of the day when data was observed (formatted as 'HH:00').
        5. 'temperature': Temperature at the time of observation (in °C).
        6. 'precipitation_category': Description of precipitation type (e.g., 'Rain', 'Snow').
        7. 'precipitation_mm': Median precipitation amount in mm.

    Errors:
    Errors encountered during the data transformation are printed to the console and an empty list is returned.
    """
    if not data:
        print('No data available to transform.')
        return None
    try:
        transformed_data = []

        # The API timestamps the forecast in UTC, so the hours to look for are
        # built in UTC and only converted to local time for the reported values.
        forecast_by_time = {
            datetime.datetime.strptime(observation['time'], '%Y-%m-%dT%H:%M:%SZ')
            .replace(tzinfo=datetime.timezone.utc): observation['data']
            for observation in data['timeSeries']}

        next_utc_hour = datetime.datetime.now(datetime.timezone.utc).replace(
            minute=0, second=0, microsecond=0) + datetime.timedelta(hours=1)
        fetched = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')

        for hours_ahead in range(24):
            utc_time = next_utc_hour + datetime.timedelta(hours=hours_ahead)
            values = forecast_by_time.get(utc_time)
            if values is None:
                continue

            local_time = utc_time.astimezone()
            date = local_time.strftime('%Y-%m-%d')
            hour = local_time.strftime('%H:00')
            temperature = values['air_temperature']
            precipitation_category_value = values['predominant_precipitation_type_at_surface']
            precipitation_category = PRECIPITATION_CATEGORIES.get(
                precipitation_category_value, 'Unknown')
            precipitation_mm = values['precipitation_amount_median']
            transformed_data.append([fetched, latitude, longitude, date,
                                    hour, temperature, precipitation_category, precipitation_mm])

        return transformed_data
    
    except KeyError as ke:
        print(f'Missing key in input data: {ke}')
        return []
    except TypeError as te:
        print(f'Invalid data type in input: {te}')
        return []
    except Exception as e:
        print(f'An unexpected error occurred: {e}')
        return []

# Load data into Excel #
def load_data_to_excel(data, location):
    if not data:
        return None
    try:
        capitalized_location = location.capitalize()

        if os.path.exists(EXCEL_FILE):
            workbook = openpyxl.load_workbook(EXCEL_FILE)
        else:
            workbook = openpyxl.Workbook()
            workbook.remove(workbook['Sheet'])

        sheet, is_new_sheet = clear_or_create_sheet(workbook, capitalized_location)
        create_excel_headers(sheet)

        for value in data: 
            sheet.append(value)
        
        print(f'\nExcel dashboard has been {"created" if is_new_sheet else f"updated with data for {capitalized_location}" }.')
        
        workbook.save(EXCEL_FILE)

    except FileNotFoundError:
        print(f'Error: The file {EXCEL_FILE} was not found.')
    except PermissionError:
        print(f'Error: Permission denied when accessing {EXCEL_FILE}.')
    except openpyxl.utils.exceptions.InvalidFileException:
        print(f'Error: The file {EXCEL_FILE} is not a valid Excel file.')
    except Exception as e:
        print(f'An unexpected error occurred: {e}')

def etl(latitude, longitude, location):
    data = extract_smhi_data(latitude, longitude)
    transformed_data = transform_smhi_data(data, latitude, longitude)
    load_data_to_excel(transformed_data, location)

# MENU: 2. Update dashboard with the new data #
def update_dashboard(excel_file_path, etl_func):
    """
    Updates the dashboard with new data for each location listed in the Excel file.

    Args:
    excel_file_path (str): Path to the Excel file.
    etl_func (function): ETL function to execute for each location.
    """
    geolocator = Nominatim(user_agent='GeocodingApp')

    try:
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet_names = workbook.sheetnames

        for location_name in sheet_names:
            try:
                location = geolocator.geocode(location_name)
                if location:
                    latitude, longitude = round(location.latitude, 6), round(location.longitude, 6)
                    etl_func(latitude, longitude, location_name)
                else:
                    print(f'Location not found for {location_name}.')
            except GeocoderTimedOut:
                print('Geocoder service timed out.')
            except GeocoderServiceError:
                print('Geocoder service error.')
            except Exception as e:
                print(f'An unexpected error occurred: {e}')
        
        print('\nDashboard updated successfully.')

    except Exception as e:
        print(f'An error occurred while updating the dashboard: {e}')
        wait_for_keypress()


